import openpyxl
import pandas as pd


def list_parse_exel(path: str, sheet_name: str, skiprows:int = 0, nrows: int = None):

    # Укажите путь к вашему Excel-файлу
    excel_file_path = path

    # Прочитайте Excel-файл
    df = pd.read_excel(excel_file_path, sheet_name=sheet_name, skiprows=skiprows, nrows=nrows) #, nrows=2)


    # for index, row in df.iterrows():
    #     print(f'Index: {index}\n'
    #           f'Row: {row}\n)'
    # print(f'{df.values}')

    # # Выведите содержимое DataFrame (df)
    # print(df.name, df.price)

    result = []
    for i in df.values:
        el = {
            "name": i[0],
            "price": i[1]
        }
        result.append(el)

    return result


def search_paint(name: str, path: str, sheet_name: str, color: str):
    column_name = 'name'
    search_value = name

    df = pd.read_excel(path, sheet_name=sheet_name)

    # Найти индекс строки, в которой находится искомое значение
    index_to_highlight = df.index[df[column_name] == search_value].tolist()

    if index_to_highlight:
        index_to_highlight = index_to_highlight[0]

        # Загрузить Excel-файл
        workbook = openpyxl.load_workbook(path)

        # Получить активный лист
        sheet = workbook.active

        # Подсветить ячейку в нужным цветом
        for col_num, value in enumerate(df.columns, start=1):
            cell = sheet.cell(row=index_to_highlight + 2, column=col_num)
            # cell.fill = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            if color == 'red':
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            elif color == 'green':
                cell.fill = openpyxl.styles.PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            elif color == 'yellow':
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')


        # Сохранить изменения
        workbook.save(path)



if __name__ == '__main__':
    print(list_parse_exel('exel/ultima.xlsx', 'Sheet1', skiprows=10, nrows=8))
    # search_paint('RAM-53NE2F', 'exel/hitachi.xlsx', 'Sheet1', 'green')



